import os
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader

current_file = os.path.abspath(__file__)
current_dir = os.path.dirname(current_file)
path_to_directory_files = os.path.join(current_dir, "resources")
path_to_directory_archive = os.path.join(current_dir, "archives")


def test_read_pdf(create_zip_for_files):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("test.pdf", "r") as pdf:
            reader = PdfReader(pdf).pages
            expected_pdf_text = "Тестовый PDF файл"
            assert expected_pdf_text in reader[0].extract_text()


def test_read_xlsx(create_zip_for_files):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("test.xlsx", "r") as xlsx:
            workbook = load_workbook(xlsx, data_only=True).active
            assert workbook.cell(row=2, column=1).value == "OU001"


def test_read_csv(create_zip_for_files):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        content = test_zip.read("test.csv")
        decoded_content = content.decode('utf-8')
        assert "Париж" in decoded_content
